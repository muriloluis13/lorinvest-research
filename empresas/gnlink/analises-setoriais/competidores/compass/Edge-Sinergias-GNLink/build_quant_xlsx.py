# -*- coding: utf-8 -*-
"""Quantificação das sinergias GNLink x Edge. Planilha viva, fase 1 (confiança alta) + TAM Bain (visão JV)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.label import DataLabelList

NAVY="15335B"; TEAL="34BBAC"; LIGHT="EAF1F4"; INP="FFF7D6"; TOT="E4E9EE"; GRAYF="F3F5F7"
thin=Side(style="thin",color="C9D4DC")
BORD=Border(left=thin,right=thin,top=thin,bottom=thin)
def F(sz=10,b=False,color="222B38"): return Font(name="Calibri",size=sz,bold=b,color=color)
def fill(c): return PatternFill("solid",fgColor=c)
def w(ws,ref,val=None,f=None,fl=None,fmt=None,al=None,bd=False,wrap=False):
    c=ws[ref]
    if val is not None: c.value=val
    c.font=f or F()
    if fl: c.fill=fill(fl)
    if fmt: c.number_format=fmt
    c.alignment=Alignment(horizontal=al or "left",vertical="center",wrap_text=wrap)
    if bd: c.border=BORD
    return c
RS='#,##0';  RS1='#,##0.0'; PCT='0%'; PCT1='0.0%'; MM='0.00'; XX='0.0"x"'

wb=openpyxl.Workbook()

# ============ PREMISSAS ============
p=wb.active; p.title="Premissas"
p.column_dimensions["A"].width=34; p.column_dimensions["B"].width=12; p.column_dimensions["C"].width=40
w(p,"A1","PREMISSAS GLOBAIS  (células amarelas = editáveis)",F(13,True,NAVY))
hdr=["Parâmetro","Valor","Unidade / nota"]
for j,h in enumerate(hdr): w(p,f"{get_column_letter(j+1)}3",h,F(10,True,"FFFFFF"),NAVY,al="center",bd=True)
rows=[("Múltiplo base",10,XX,"x EBITDA, base"),
      ("Múltiplo teto",12,XX,"x EBITDA, teto"),
      ("Desconto",0.12,PCT,"a.a."),
      ("Margem off-grid (baixa)",0.80,'0.00',"R$/m³"),
      ("Margem off-grid (alta)",0.92,'0.00',"R$/m³"),
      ("Alíquota IR/CSLL",0.34,PCT,""),
      ("Dias por ano",365,'0',""),
      ("Captura do TAM (baixa)",0.05,PCT,"% do mercado endereçável"),
      ("Captura do TAM (alta)",0.10,PCT,"% do mercado endereçável")]
r=4
for lab,val,fmt,note in rows:
    w(p,f"A{r}",lab,F(),bd=True); w(p,f"B{r}",val,F(10,True),INP,fmt,"center",bd=True); w(p,f"C{r}",note,F(9,color="6B7A8C"),bd=True); r+=1
MB="Premissas!$B$4"; MT="Premissas!$B$5"; DE="Premissas!$B$6"; MGL="Premissas!$B$7"; MGH="Premissas!$B$8"; IR="Premissas!$B$9"; CPL="Premissas!$B$11"; CPH="Premissas!$B$12"

# ============ TAM BAIN ============
t=wb.create_sheet("TAM_Bain")
for col,wd in zip("ABCDEFGHI",[30,12,7,7,7,7,7,13,13]): t.column_dimensions[col].width=wd
w(t,"A1","MERCADO POTENCIAL OFF-GRID  ·  Bain (visão JV, jul/2020)  ·  13,0 MMm³/d",F(13,True,NAVY))
w(t,"A2","Sizing bottom-up do mercado brasileiro de substituição por caso de uso e região (fonte independente).",F(9,color="6B7A8C"))
H=["Caso de uso","TAM (MMm³/d)","N","NE","SE","S","CO","N+NE","N+NE+S"]
for j,h in enumerate(H): w(t,f"{get_column_letter(j+1)}4",h,F(10,True,"FFFFFF"),NAVY,al="center",bd=True)
data=[("Óleo combustível",5.2,.42,.34,.12,.08,.04),
      ("GLP industrial",3.0,.03,.19,.45,.24,.09),
      ("CDL (GLP residencial)",2.1,.23,.23,.27,.26,.01),
      ("Postos GNV",0.4,.01,.14,.60,.19,.06),
      ("Frota BR",0.3,.03,.18,.47,.15,.15),
      ("Transportadoras/embarcadores",2.0,None,None,None,None,None)]
r=5
for nm,tam,n,ne,se,su,co in data:
    w(t,f"A{r}",nm,F(),bd=True); w(t,f"B{r}",tam,F(10,True),None,MM,"center",bd=True)
    for j,v in zip("CDEFG",[n,ne,se,su,co]):
        w(t,f"{j}{r}",v,F(),None,PCT,"center",bd=True)
    if n is not None:
        w(t,f"H{r}",f"=B{r}*(C{r}+D{r})",F(),None,MM,"center",bd=True)
        w(t,f"I{r}",f"=B{r}*(C{r}+D{r}+F{r})",F(),None,MM,"center",bd=True)
    else:
        w(t,f"H{r}","n/d",F(9,color="6B7A8C"),None,None,"center",bd=True); w(t,f"I{r}","n/d",F(9,color="6B7A8C"),None,None,"center",bd=True)
    r+=1
tot=r
w(t,f"A{tot}","Total",F(10,True),TOT,bd=True); w(t,f"B{tot}",f"=SUM(B5:B{r-1})",F(10,True),TOT,MM,"center",bd=True)
for j in "CDEFG": w(t,f"{j}{tot}",None,None,TOT,bd=True)
w(t,f"H{tot}",f"=SUM(H5:H9)",F(10,True),TOT,MM,"center",bd=True); w(t,f"I{tot}",f"=SUM(I5:I9)",F(10,True),TOT,MM,"center",bd=True)
w(t,f"A{tot+2}","N+NE = geografia exclusiva da GNLink (fora do raio de Santos); Sul = GNLink forte, parcialmente contestável.",F(9,color="6B7A8C"))
w(t,f"A{tot+3}","Transportadoras sem split regional publicado (excluídas do endereçável N+NE, critério conservador).",F(9,color="6B7A8C"))
w(t,f"A{tot+4}","Óleo combustível (maior balde, 5,2) é 76% N+NE: o maior mercado está na geografia que Santos não alcança.",F(9,True,TEAL))
TAM_NNE=f"TAM_Bain!$H${tot}"; TAM_NNES=f"TAM_Bain!$I${tot}"
TAM_OC="TAM_Bain!$B$5"; TAM_GLP="TAM_Bain!$B$6"; TAM_GNV="TAM_Bain!$B$9"; TAM_FBR="TAM_Bain!$B$8"; TAM_TR="TAM_Bain!$B$10"

# ============ CALC_B (confiança alta) ============
b=wb.create_sheet("Calc_B")
b.column_dimensions["A"].width=42; b.column_dimensions["B"].width=13; b.column_dimensions["C"].width=13; b.column_dimensions["D"].width=40
def sec(ws,row,title):
    w(ws,f"A{row}",title,F(11,True,"FFFFFF"),NAVY,bd=True)
    for col in "BCD": w(ws,f"{col}{row}",None,None,NAVY,bd=True)
def line(ws,row,lab,val,fmt=None,inp=False,note="",bold=False,out=False):
    w(ws,f"A{row}",lab,F(10,bold),TOT if out else None,bd=True)
    w(ws,f"B{row}",val,F(10,True if(inp or bold or out)else False),INP if inp else(TOT if out else None),fmt,"center",bd=True)
    w(ws,f"D{row}",note,F(9,color="6B7A8C"),bd=True)

w(b,"A1","QUANTIFICAÇÃO — SINERGIAS INCREMENTAIS DE CONFIANÇA ALTA (Grupo II) e GRUPO I",F(13,True,NAVY))
# B7 Regás
sec(b,3,"B7 · Fábrica de regás in-house — capex de importação evitado  (Grupo II · incremental)")
line(b,4,"Capex Fase 2 da Edge (R$ mi)",1500,RS,inp=True)
line(b,5,"Fatia da regás no capex off-grid",0.16,PCT,inp=True)
line(b,6,"Capex de regás na Fase 2 (custo importado)","=B4*B5",RS,note="R$ mi")
line(b,7,"Custo GNLink / custo importado",0.333,PCT,inp=True,note="regás GNLink ~3x mais barata")
line(b,8,"Capex de regás pela GNLink","=B6*B7",RS)
line(b,9,"Economia bruta de capex","=B6-B8",RS,bold=True)
line(b,10,"Fator de VP (spend 2027-29, ~2 anos)","=1/(1+"+DE+")^2",'0.000')
line(b,11,"Capex evitado — VP (baixa)","=B9*B10",RS)
line(b,12,"Capex evitado — nominal (alta)","=B9",RS)
w(b,"A13","Aceleração de rampa: CORTADA — não quantificável de forma defensável (a conta atribuía ~metade do salto de crescimento da Edge ao timing da regás).",F(9,True,"B9433A"),wrap=True); b.row_dimensions[13].height=30
line(b,19,"TOTAL B7 — baixa (só capex evitado, VP)","=B11",RS,out=True)
line(b,20,"TOTAL B7 — alta (só capex evitado, nominal)","=B12",RS,out=True)
B7L="Calc_B!$B$19"; B7H="Calc_B!$B$20"
# B6 Refi
sec(b,22,"B6 · Refinanciamento / custo da dívida  (CORTADA — abaixo do EBITDA, não cabe em EV/EBITDA)")
line(b,23,"Dívida bruta GNLink (R$ mi)",330,RS,inp=True,note="corrigido para R$330 mi; bloco mantido só para referência, NÃO entra no total")
line(b,24,"Redução de custo — baixa (bps)",0.02,PCT,inp=True,note="200 bps")
line(b,25,"Redução de custo — alta (bps)",0.035,PCT,inp=True,note="350 bps")
line(b,26,"Economia anual — baixa","=B23*B24",RS,note="R$ mi/ano")
line(b,27,"Economia anual — alta","=B23*B25",RS,note="R$ mi/ano")
line(b,28,"TOTAL B6 — baixa (x múltiplo base)","=B26*"+MB,RS,out=True)
line(b,29,"TOTAL B6 — alta (x múltiplo base)","=B27*"+MB,RS,out=True)
B6L="Calc_B!$B$28"; B6H="Calc_B!$B$29"
# B10 Escudo fiscal
sec(b,31,"B10 · Escudo fiscal de prejuízos acumulados  (Grupo II · A PREENCHER)")
line(b,32,"Prejuízos fiscais acumulados (R$ mi)",162,RS,inp=True,note="NOL informado pelo usuário")
line(b,33,"Escudo nominal (x alíquota)","=B32*"+IR,RS)
line(b,34,"Fator de aproveitamento / VP",1.0,PCT,inp=True,note="100% (bruto, sem haircut de trava 30%/ano nem desconto)")
line(b,35,"TOTAL B10","=B33*B34",RS,out=True,note="fica R$0 até preencher o NOL")
B10V="Calc_B!$B$35"
# B11 Argentina (Grupo I)
sec(b,37,"B11 · Argentina, capex do supridor  (Grupo I · JÁ no valuation R$1,2-1,4 bn)")
line(b,38,"EBITDA Argentina maduro (R$ mi)",100,RS,inp=True,note="F1 R$50 (2028) + F2 R$50 (2030)")
line(b,39,"Anos de desconto (média)",3,'0',inp=True)
line(b,40,"Valor no valuation — 10x (desc)","=B38*"+MB+"/(1+"+DE+")^B39",RS,out=True)
line(b,41,"Valor no valuation — 12x (desc)","=B38*"+MT+"/(1+"+DE+")^B39",RS,out=True)
B11L="Calc_B!$B$40"; B11H="Calc_B!$B$41"
# B12 Eneva (Grupo I)
sec(b,43,"B12 · Projeto Eneva  (Grupo I · JÁ no valuation)")
line(b,44,"EBITDA Eneva (R$ mi)",50,RS,inp=True,note="take assinado, 2027")
line(b,45,"Anos de desconto",1,'0',inp=True)
line(b,46,"Valor no valuation — 10x (desc)","=B44*"+MB+"/(1+"+DE+")^B45",RS,out=True)
line(b,47,"Valor no valuation — 12x (desc)","=B44*"+MT+"/(1+"+DE+")^B45",RS,out=True)
B12L="Calc_B!$B$46"; B12H="Calc_B!$B$47"
# --- FASE 2 (ordem de grandeza) ---
sec(b,49,"B1 · Crescimento das plantas (RN + Itabuna)  (Grupo I · eleva a base do valuation · o.g.)")
line(b,50,"Volume incremental RN (mil m³/d)",40,RS,inp=True,note="ordem de grandeza")
line(b,51,"Volume incremental Itabuna/BA (mil m³/d)",40,RS,inp=True,note="ordem de grandeza")
line(b,52,"Volume incremental total (mil m³/d)","=B50+B51",RS)
line(b,53,"EBITDA incremental/ano — baixa (R$ mi)","=B52*"+MGL+"*0.365",RS)
line(b,54,"EBITDA incremental/ano — alta (R$ mi)","=B52*"+MGH+"*0.365",RS)
line(b,55,"Valor — baixa (10x)","=B53*"+MB,RS,out=True)
line(b,56,"Valor — alta (12x)","=B54*"+MT,RS,out=True)
B1L="Calc_B!$B$55"; B1H="Calc_B!$B$56"
sec(b,58,"B3 · Molécula, arbitragem de despacho  (Grupo II · incremental · o.g.)")
line(b,59,"Volume da combinação (MMm³/d)",0.69,MM,inp=True,note="curva cheia GNLink")
line(b,60,"Fração acionável",0.20,PCT,inp=True)
line(b,61,"Economia por m³ — baixa (R$/m³)",0.05,'0.00',inp=True)
line(b,62,"Economia por m³ — alta (R$/m³)",0.15,'0.00',inp=True)
line(b,63,"Economia anual — baixa (R$ mi)","=B59*B60*B61*365",RS)
line(b,64,"Economia anual — alta (R$ mi)","=B59*B60*B62*365",RS)
line(b,65,"Valor — baixa (10x)","=B63*"+MB,RS,out=True)
line(b,66,"Valor — alta (10x)","=B64*"+MB,RS,out=True)
B3L="Calc_B!$B$65"; B3H="Calc_B!$B$66"
sec(b,68,"B5 · Logística: origem mais próxima + backhaul na rede combinada  (Grupo II · incremental · o.g.)")
line(b,69,"Volume beneficiado (MMm³/d)",0.28,MM,inp=True,note="volume GNLink que ganha com a rede combinada")
line(b,70,"Economia logística/m³ — baixa (R$/m³)",0.05,'0.00',inp=True,note="fração da margem ~R$0,80/m³")
line(b,71,"Economia logística/m³ — alta (R$/m³)",0.12,'0.00',inp=True)
line(b,72,"Economia anual — baixa (R$ mi)","=B69*B70*365",RS)
line(b,73,"Economia anual — alta (R$ mi)","=B69*B71*365",RS)
line(b,74,"Valor — baixa (10x)","=B72*"+MB,RS,out=True)
line(b,75,"Valor — alta (10x)","=B73*"+MB,RS,out=True)
B5L="Calc_B!$B$74"; B5H="Calc_B!$B$75"
w(b,"A76","Narrativa: 3 plantas (NE/S) + o terminal de Santos permitem suprir cada cliente pela origem mais próxima e recarregar a carreta no nó mais próximo, em vez de voltar vazia à planta de origem. Corta km ocioso e distância média de suprimento; é custo acima do EBITDA, logo cabe no EV/EBITDA.",F(9,color="4A5568"),wrap=True); b.row_dimensions[76].height=42
sec(b,77,"B9 · Aceleração regulatória (licenciamento)  (CORTADA — sem volume real travado por licença hoje)")
line(b,78,"Volume travado por licença (mil m³/d)",0,RS,inp=True,note="hoje não há volume travado; bloco mantido só para referência")
line(b,79,"Meses de aceleração — baixa",6,'0',inp=True)
line(b,80,"Meses de aceleração — alta",12,'0',inp=True)
line(b,81,"EBITDA anual do volume travado (R$ mi)","=B78*"+MGL+"*0.365",RS)
line(b,82,"Valor da aceleração — baixa","=B81*B79/12",RS,out=True,note="antecipação, não perpétuo")
line(b,83,"Valor da aceleração — alta","=B78*"+MGH+"*0.365*B80/12",RS,out=True)
B9L="Calc_B!$B$82"; B9H="Calc_B!$B$83"
sec(b,85,"A10 · Backup cruzado  (Grupo II · incremental · o.g.)")
line(b,86,"Volume em risco por parada (mil m³/d)",93,RS,inp=True,note="~uma planta")
line(b,87,"Eventos por ano",1,'0',inp=True)
line(b,88,"Duração — baixa (dias)",5,'0',inp=True)
line(b,89,"Duração — alta (dias)",15,'0',inp=True)
line(b,90,"Perda evitada/ano — baixa (R$ mi)","=B86*B88*B87*"+MGL+"/1000",RS)
line(b,91,"Perda evitada/ano — alta (R$ mi)","=B86*B89*B87*"+MGH+"/1000",RS)
line(b,92,"Valor — baixa (10x)","=B90*"+MB,RS,out=True)
line(b,93,"Valor — alta (10x)","=B91*"+MB,RS,out=True)
A10L="Calc_B!$B$92"; A10H="Calc_B!$B$93"
sec(b,95,"B8 · Redes locais nas distribuidoras Compass  (Grupo III · lente alternativa · o.g.)")
line(b,96,"Potencial redes locais (MMm³/d)",0.916,MM,inp=True,note="concessões Compass: Sul 433k + NE")
line(b,97,"Captura — baixa",0.10,PCT,inp=True)
line(b,98,"Captura — alta",0.20,PCT,inp=True)
line(b,99,"EBITDA/ano — baixa (R$ mi)","=B96*B97*"+MGL+"*365",RS)
line(b,100,"EBITDA/ano — alta (R$ mi)","=B96*B98*"+MGH+"*365",RS)
line(b,101,"Valor — baixa (10x)","=B99*"+MB,RS,out=True)
line(b,102,"Valor — alta (12x)","=B100*"+MT,RS,out=True)
B8L="Calc_B!$B$101"; B8H="Calc_B!$B$102"

# ============ CALC_A (TAM Bain) ============
a=wb.create_sheet("Calc_A")
a.column_dimensions["A"].width=42; a.column_dimensions["B"].width=13; a.column_dimensions["C"].width=13; a.column_dimensions["D"].width=40
w(a,"A1","DIMENSÃO DE MERCADO — BLOCO A (atributos), ancorado no TAM da Bain",F(13,True,NAVY))
w(a,"A2","Tamanho da oportunidade, não valor comprometido. Overlap com o pipeline próprio da GNLink e com B8 (redes locais/CDL): não somar.",F(9,color="6B7A8C"))
# A1
sec(a,4,"A1 · Extensão geográfica  (Grupo IV · dimensão da oportunidade)")
line(a,5,"Mercado endereçável N+NE (MMm³/d)","="+TAM_NNE,MM,note="exclusivo GNLink, do TAM Bain")
line(a,6,"Mercado endereçável N+NE+Sul (MMm³/d)","="+TAM_NNES,MM)
line(a,7,"Volume capturado — baixa (MMm³/d)","=B5*"+CPL,MM)
line(a,8,"Volume capturado — alta (MMm³/d)","=B6*"+CPH,MM)
line(a,9,"EBITDA/ano — baixa (R$ mi)","=B7*"+MGL+"*"+"Premissas!$B$10",RS,note="vol(MM) x margem x dias")
line(a,10,"EBITDA/ano — alta (R$ mi)","=B8*"+MGH+"*Premissas!$B$10",RS)
line(a,11,"Valor indicativo — baixa (10x)","=B9*"+MB,RS,out=True)
line(a,12,"Valor indicativo — alta (12x)","=B10*"+MT,RS,out=True)
# A9
sec(a,14,"A9 · Expertise com transporte  (Grupo IV · opção / TAM)")
line(a,15,"TAM transporte (MMm³/d)","="+TAM_TR+"+"+TAM_FBR+"+"+TAM_GNV,MM,note="transportadoras + Frota BR + GNV")
line(a,16,"Volume capturado — baixa","=B15*"+CPL,MM)
line(a,17,"EBITDA/ano — baixa (R$ mi)","=B16*"+MGL+"*Premissas!$B$10",RS)
line(a,18,"Valor indicativo — baixa (10x)","=B17*"+MB,RS,out=True)
# A2
sec(a,20,"A2 · Diversidade de produto / expertise comercial  (Grupo IV · TAM de contexto)")
line(a,21,"TAM GLP + OC (MMm³/d)","="+TAM_OC+"+"+TAM_GLP,MM,note="mercado dominado pela expertise (81% do pipeline)")
w(a,"A22","GNC e infraestrutura ampliam o alcance ao cliente pequeno (segmento que a Edge não atende). Valor via captura em A1.",F(9,color="6B7A8C"))

# ============ RESUMO ============
s=wb.create_sheet("Resumo"); wb.move_sheet("Resumo",-(len(wb.sheetnames)-1))
for col,wd in zip("ABCDEFGH",[7,34,9,9,13,13,30,20]): s.column_dimensions[col].width=wd
w(s,"A1","QUANTIFICAÇÃO DAS SINERGIAS  ·  GNLink × Edge",F(14,True,NAVY))
w(s,"A2","Regra anti-double-count: Grupo I já no valuation (R$1,2-1,4 bn) · Grupo II soma sobre ele · Grupo III é lente alternativa · Grupo IV é dimensão da oportunidade.",F(9,color="6B7A8C"),wrap=True)
H=["ID","Sinergia","Grupo","Confiança","R$ mi (baixa)","R$ mi (alta)","Método / base","Status"]
for j,h in enumerate(H): w(s,f"{get_column_letter(j+1)}4",h,F(10,True,"FFFFFF"),NAVY,al="center",bd=True)
# (id, nome, grupo, conf, baixa, alta, metodo, status)
rowsR=[
("A1","Extensão geográfica","IV","média","=Calc_A!$B$11","=Calc_A!$B$12","TAM Bain N+NE x captura x margem","Dimensão (não somar)"),
("A2","Diversidade de produto","IV","proxy","","","TAM GLP+OC = 8,2 MMm³/d","Contexto"),
("A3","Tamanho de cliente","IV","—","","","enabler (dentro de A2)","Qualitativo"),
("A4","Tipos de cliente","IV","—","","","enabler (B8/A9)","Qualitativo"),
("A5","Segmentos industriais","IV","—","","","enabler","Qualitativo"),
("A6","Expertise comercial","IV","—","","","enabler (valor em A1/A7)","Qualitativo"),
("A7","Execução comprovada","I+III","média","","","aceleração/de-risk da rampa","Fase 2"),
("A8","Expertise de engenharia","IV","—","","","enabler (valor em B7)","Qualitativo"),
("A9","Expertise com transporte","IV","baixa","=Calc_A!$B$18","","TAM transporte 2,7 MMm³/d x captura","Dimensão (não somar)"),
("A10","Backup e resiliência","II","baixa","="+A10L,"="+A10H,"perda esperada evitada (o.g.)","Quantificado (o.g.)"),
("B1","Crescimento das plantas","I","média","="+B1L,"="+B1H,"EBITDA expansões RN+Itabuna (o.g.)","Eleva a base"),
("B2","Redução do risco de capex","III","baixa","","","de-risk do plano (framing)","Qualitativo"),
("B3","Molécula, arbitragem","II","média","="+B3L,"="+B3H,"economia por m³ x volume (o.g.)","Quantificado (o.g.)"),
("B4","Molécula própria (Tradener)","IV","baixa","","","opção (a avaliar)","Falta dado"),
("B5","Otimização logística","II","baixa","="+B5L,"="+B5H,"economia logística/m³ (o.g.)","Quantificado (o.g.)"),
("B6","Refinanciamento","—","cortada","","","abaixo do EBITDA; não cabe em EV/EBITDA","Cortada"),
("B7","Fábrica de regás","II","ALTA","="+B7L,"="+B7H,"capex de importação evitado (VP)","Quantificado"),
("B8","Redes locais","III","média-baixa","="+B8L,"="+B8H,"916 mil m³/d x captura x margem (o.g.)","Lente alternativa (o.g.)"),
("B9","Aceleração regulatória","—","cortada","","","sem volume real travado por licença hoje","Cortada"),
("B10","Equity / escudo fiscal","II","média","="+B10V,"="+B10V,"NOL R$162 mi x 34% (valor de face, sem múltiplo)","Quantificado"),
("B11","Argentina","I","ALTA","="+B11L,"="+B11H,"EBITDA R$100 mi x múltiplo","Já no valuation"),
("B12","Eneva","I","ALTA","="+B12L,"="+B12H,"EBITDA R$50 mi x múltiplo","Já no valuation"),
]
r=5
for idd,nm,gr,cf,lo,hi,me,st in rowsR:
    w(s,f"A{r}",idd,F(10,True,NAVY),bd=True); w(s,f"B{r}",nm,F(),bd=True)
    w(s,f"C{r}",gr,F(),None,None,"center",bd=True); w(s,f"D{r}",cf,F(10,True if cf=="ALTA" else False,TEAL if cf=="ALTA" else "222B38"),None,None,"center",bd=True)
    w(s,f"E{r}",lo if lo!="" else "—",F(),None,RS,"center",bd=True); w(s,f"F{r}",hi if hi!="" else "—",F(),None,RS,"center",bd=True)
    w(s,f"G{r}",me,F(9,color="4A5568"),bd=True,wrap=True); w(s,f"H{r}",st,F(9,color="6B7A8C"),bd=True)
    r+=1
# Consolidação
cc=r+1
GII_L="="+"+".join([B7L,B5L,B10V,B3L,A10L]); GII_H="="+"+".join([B7H,B5H,B10V,B3H,A10H])
w(s,f"A{cc}","CONSOLIDAÇÃO",F(12,True,NAVY))
w(s,f"A{cc+1}","Grupo II · sinergias incrementais da combinação (SOMAM)",F(10,True),None,None,"left")
w(s,f"E{cc+1}",GII_L,F(10,True),TOT,RS,"center",bd=True); w(s,f"F{cc+1}",GII_H,F(10,True),TOT,RS,"center",bd=True)
w(s,f"A{cc+2}","regás (capex evitado) + logística + escudo fiscal (NOL R$162 mi) + arbitragem + backup. Cortadas: refi (abaixo do EBITDA), regulatório (sem volume travado) e a aceleração da regás.",F(9,color="6B7A8C"),wrap=True); s.row_dimensions[cc+2].height=30
w(s,f"A{cc+3}","Grupo III · lente alternativa (NÃO somar com o Grupo II)",F(10,True),None,None,"left")
w(s,f"A{cc+4}","   deslocamento no valuation da Edge",F(10)); w(s,f"E{cc+4}",973,F(10),None,RS,"center",bd=True); w(s,f"G{cc+4}","~R$1 bn (via régua BTG)",F(9,color="6B7A8C"))
w(s,f"A{cc+5}","   redes locais nas distribuidoras Compass",F(10)); w(s,f"E{cc+5}","="+B8L,F(10),None,RS,"center",bd=True); w(s,f"F{cc+5}","="+B8H,F(10),None,RS,"center",bd=True); w(s,f"G{cc+5}","pocket diferente da Edge",F(9,color="6B7A8C"))
w(s,f"A{cc+6}","Grupo I · já no valuation da GNLink (R$1,2-1,4 bn)",F(10,True),None,None,"left")
w(s,f"A{cc+7}","   crescimento das plantas + Argentina + Eneva",F(10)); w(s,f"E{cc+7}","="+"+".join([B1L,B11L,B12L]),F(10),None,RS,"center",bd=True); w(s,f"F{cc+7}","="+"+".join([B1H,B11H,B12H]),F(10),None,RS,"center",bd=True)
w(s,f"A{cc+9}","Leitura: o Grupo II (incrementais) soma SOBRE a GNLink; as duas lentes de valor (P&L GNLink vs. deslocamento na Edge) devem ficar na mesma ordem de grandeza, não somadas. Fase 2 é ordem de grandeza (o.g.) e depende das premissas editáveis.",F(9,True,color="4A5568"),wrap=True)
s.row_dimensions[cc+9].height=42

# ============ WATERFALL ============
wf=wb.create_sheet("Waterfall")
for col,wd in zip("ABCD",[26,13,13,13]): wf.column_dimensions[col].width=wd
w(wf,"A1","SINERGIAS INCREMENTAIS DA COMBINAÇÃO  ·  Grupo II (ponto médio das faixas)",F(13,True,NAVY))
w(wf,"A3","Apenas as sinergias que somam sobre o valor da GNLink; sem o valuation-base.",F(9,color="6B7A8C"))
for j,h in enumerate(["Etapa","Base (inv.)","Valor","Acumulado"]): w(wf,f"{get_column_letter(j+1)}5",h,F(10,True,"FFFFFF"),NAVY,al="center",bd=True)
steps=[("Regás (capex evitado)","=(Calc_B!$B$19+Calc_B!$B$20)/2"),
       ("+ Logística","=(Calc_B!$B$74+Calc_B!$B$75)/2"),
       ("+ Escudo fiscal","=Calc_B!$B$35"),
       ("+ Arbitragem molécula","=(Calc_B!$B$65+Calc_B!$B$66)/2"),
       ("+ Backup","=(Calc_B!$B$92+Calc_B!$B$93)/2")]
r=6
for i,(lab,val) in enumerate(steps):
    w(wf,f"A{r}",lab,F(10,i==0),bd=True)
    w(wf,f"B{r}",0 if i==0 else f"=D{r-1}",F(),None,RS,"center",bd=True)
    w(wf,f"C{r}",val,F(10,i==0),None,RS,"center",bd=True)
    w(wf,f"D{r}",f"=B{r}+C{r}",F(),None,RS,"center",bd=True)
    r+=1
w(wf,f"A{r}","Total das sinergias (Grupo II)",F(10,True),TOT,bd=True)
w(wf,f"B{r}",0,F(10,True),TOT,RS,"center",bd=True)
w(wf,f"C{r}",f"=D{r-1}",F(10,True),TOT,RS,"center",bd=True)
w(wf,f"D{r}",f"=C{r}",F(10,True),TOT,RS,"center",bd=True)
last=r
ch=BarChart(); ch.type="col"; ch.grouping="stacked"; ch.overlap=100; ch.gapWidth=45
ch.add_data(Reference(wf,min_col=2,max_col=3,min_row=5,max_row=last),titles_from_data=True)
ch.set_categories(Reference(wf,min_col=1,min_row=6,max_row=last))
ch.series[0].graphicalProperties=GraphicalProperties(); ch.series[0].graphicalProperties.noFill=True
ch.series[1].graphicalProperties=GraphicalProperties(solidFill="34BBAC")
ch.series[1].dLbls=DataLabelList(); ch.series[1].dLbls.showVal=True
ch.title="Sinergias incrementais, Grupo II (R$ mi)"; ch.legend=None; ch.y_axis.title="R$ mi"
ch.x_axis.delete=False; ch.y_axis.delete=False; ch.height=10.5; ch.width=23
wf.add_chart(ch,"F5")
w(wf,f"A{last+2}","Lente alternativa (NÃO somar com esta ponte): deslocamento no valuation da Edge ~R$973 mi + redes locais nas distribuidoras da Compass. É a mesma criação de valor vista pelo outro lado.",F(9,True,color="4A5568"),wrap=True)
wf.row_dimensions[last+2].height=42
wb.move_sheet("Waterfall",-(len(wb.sheetnames)-2))

wb.save("Quantificacao-Sinergias-GNLink-Edge.xlsx")
print("OK salvo.")
# sanity (python-side, não são as fórmulas do Excel)
b7l=160*(1/1.12**2)
B3l,B3h=0.69*0.20*0.05*365*10,0.69*0.20*0.15*365*10
B5l,B5h=0.28*0.05*365*10,0.28*0.12*365*10
A10l,A10h=93*5*1*0.80/1000*10,93*15*1*0.92/1000*10
B10v=162*0.34*1.0
GIIl=b7l+B5l+B10v+B3l+A10l; GIIh=160+B5h+B10v+B3h+A10h
print(f"B7 {b7l:.0f}-160 | B5 {B5l:.0f}-{B5h:.0f} | B10 {B10v:.0f} | B3 {B3l:.0f}-{B3h:.0f} | A10 {A10l:.0f}-{A10h:.0f}")
print(f"Grupo II TOTAL (pós-cortes) ~ R${GIIl:.0f}-{GIIh:.0f} mi | midpoint ~R${(GIIl+GIIh)/2:.0f} mi")
